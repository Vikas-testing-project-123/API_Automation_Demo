import requests
import json
import jsonpath
import openpyxl

url = "http://bam.kockpit.in:4001/userLogin"
def update_content(jsonBody, param, parameter_value):
    return jsonBody.replace("%%"+ param + "%%", parameter_value)

def Read_Excel(excelPath, SheetName, Scenario):
    workBook = openpyxl.load_workbook(excelPath)
    Sheet=workBook.get_sheet_by_name(name=SheetName)
    # Sheet = workBook.active
    Cell_Obj = Sheet.cell(row= 1, column= 1)

    max_row=Sheet.max_row
    Scenario_Data = []
    if Cell_Obj.value=="Scenario":
        i = 2
        columni=1
        while i<=max_row:
            if Scenario == Sheet.cell(row=i,column=columni).value:
                max_column=Sheet.max_column
                while columni<max_column:
                    columni+=1
                    Scenario_Data.append(str(Sheet.cell(i,columni).value))
            i+=1
    workBook.close()
    return Scenario_Data

def write_Excel(excelPath, SheetName, Scenario,status):
    workBook = openpyxl.load_workbook(excelPath)
    Sheet=workBook.get_sheet_by_name(name=SheetName)
    Cell_Obj = Sheet.cell(row= 1, column= 1)

    max_row=Sheet.max_row
    Scenario_Data = []
    if Cell_Obj.value=="Scenario":
        i = 2
        columni=1
        while i<=max_row:
            if Scenario == Sheet.cell(row=i,column=columni).value:
                max_column=Sheet.max_column
                c1=Sheet.cell(i,max_column)
                c1.value=status
            i+=1
    workBook.save(excelPath)


excelPath="C:\\Users\\TA0134\\PycharmProjects\\API_Testing\\TestData\\UserLogin.xlsx"
testcases=["All valid parameter","Blank UserId","Blank Company Domain","Blank Password","All Blank","Invalid Company Domain","Invalid UserId","Invalid Password","All Invalid"]
testcases1=["All valid parameter"]
for test in testcases:
    print("*************************"+test+"******************************")
    testData=Read_Excel(excelPath, "User_Login", test)
    print(testData)
    i=0
    while i< len(testData):
       if testData[i] == "Blank":
           testData[i]=testData[i].replace("Blank","")
       i+=1

    print("Final TestData"+str(testData))
    file = open('C:\\Users\\TA0134\\PycharmProjects\\API_Testing\\Json_files\\User_Login.json', 'r')  # open the file in read only mode
    json_input = file.read()# It is in string format so need to convert in json
    # Update the data in the json file
    json_input=update_content(json_input, "CompanyDomain", testData[0])
    json_input=update_content(json_input, "UserId", testData[1])
    json_input=update_content(json_input, "Password", testData[2])
    request_json = json.loads(json_input) # json.loads we use to convert in json format

    # Make the post request with the json input
    response = requests.post(url, request_json)
    print(response.status_code)

    if response.status_code==int(testData[3]):
        write_Excel(excelPath,"User_Login",test,"Pass")
    else:
        write_Excel(excelPath,"User_Login",test,"Fail")

    # Get the values of the json file provided and validate the fields

    # assert response.status_code == 200
    print(response.text) #Content on the status conde
    # print(response.headers.get("X-Powered-By")) # print a specific key value from the headers.
    print("*************************Test End******************************")
