import requests
import json
import openpyxl

# Read_Excel method is used to read a excel file
def Read_Excel(excelPath, SheetName, Scenario):
        workBook = openpyxl.load_workbook(excelPath)
        Sheet = workBook.get_sheet_by_name(name=SheetName)
        Cell_Obj = Sheet.cell(row=1, column=1)

        max_row = Sheet.max_row
        Scenario_Data = []
        if Cell_Obj.value == "Scenario":
            i = 2
            columni = 1
            while i <= max_row:
                if Scenario == Sheet.cell(row=i, column=columni).value:
                    max_column = Sheet.max_column
                    while columni < max_column:
                        columni += 1
                        Scenario_Data.append(str(Sheet.cell(i, columni).value))
                i += 1
        workBook.close()
        return Scenario_Data

# update_content method is used for updating the json data
def update_content(jsonBody, param, parameter_value):
    return jsonBody.replace("%%" + param + "%%", parameter_value)

# write_Excel method is used to write testcase status to actual excel
def write_Excel(excelPath, SheetName, Scenario, status):
    workBook = openpyxl.load_workbook(excelPath)
    Sheet=workBook.get_sheet_by_name(name=SheetName)
    Cell_Obj = Sheet.cell(row=1, column=1)

    max_row = Sheet.max_row
    Scenario_Data = []
    if Cell_Obj.value == "Scenario":
        i = 2
        columni = 1
        while i <= max_row:
            if Scenario == Sheet.cell(row=i, column=columni).value:
                max_column = Sheet.max_column
                c1 = Sheet.cell(i, max_column)
                c1.value = status
            i += 1
    workBook.save(excelPath)

def test_dataSales():
    url = "http://bam.kockpit.in:4001/data/Sales"
    excelPath = "C:\\Users\\TA0134\\PycharmProjects\\API_Testing\\TestData\\UserLogin.xlsx"
    testcases = ["All valid parameter", "Blank UserId", "Blank Level", "Blank Type", "All Blank",
                 "All Invalid", "Invalid UserId", "Invalid Level", "Invalid Type", "Blank fiscal year",
                 "Invalid fiscal year", "Blank CompanyDomain", "Invalid CompanyDomain"]
    testcases1 = ["All valid parameter"]
    for test in testcases:
        print("*************************" + test + "******************************")
        testData = Read_Excel(excelPath, "dataSales", test)
        i = 0
        while i < len(testData):
            if testData[i] == "Blank":
                testData[i] = testData[i].replace("Blank", "")
            i += 1

        print("Final TestData" + str(testData))
        file = open('C:\\Users\\TA0134\\PycharmProjects\\API_Testing\\Json_files\\dataSales.json', 'r')  # open the file in read only mode
        json_input = file.read()  # It is in string format so need to convert in json
        # Update the data in the json file
        json_input = update_content(json_input, "CompanyDomain", testData[0])
        json_input = update_content(json_input, "UserId", testData[1])
        json_input = update_content(json_input, "Level", testData[2])
        json_input = update_content(json_input, "Type", testData[3])
        json_input = update_content(json_input, "FiscalYear", testData[4])
        request_json = json.loads(json_input)  # json.loads we use to convert in json format

        # Make the post request with the json input
        response = requests.post(url, request_json)
        print(response.status_code)
        #assert response.status_code == int(testData[5])
        if response.status_code == int(testData[5]):
            write_Excel(excelPath, "dataSales", test, "Pass:- "+response.text)
        else:
            write_Excel(excelPath, "dataSales", test, "Fail-"+response.text)

        # Get the values of the json file provided and validate the fields

        # assert response.status_code == 200
        print(response.text)  # Content on the status conde
        # print(response.headers.get("X-Powered-By")) # print a specific key value from the headers.
        print("*************************Test End******************************")
