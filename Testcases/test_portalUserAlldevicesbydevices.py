import requests
import json
import openpyxl
import baseUrlHeader

# Read_Excel method is used to read a excel file
def Read_Excel(excelPath, SheetName, Scenario):
        workBook = openpyxl.load_workbook(excelPath)
        Sheet = workBook.get_sheet_by_name(name=SheetName)
        Cell_Obj = Sheet.cell(row=1, column=1)

        max_row = Sheet.max_row
        Scenario_Data = []
        if Cell_Obj.value == "Scenario":
            i = 2
            columni = 1
            while i <= max_row:
                if Scenario == Sheet.cell(row=i, column=columni).value:
                    max_column = Sheet.max_column
                    while columni < max_column:
                        columni += 1
                        Scenario_Data.append(str(Sheet.cell(i, columni).value))
                i += 1
        workBook.close()
        return Scenario_Data

# update_content method is used for updating the json data
def update_content(jsonBody, param, parameter_value):
    return jsonBody.replace("%%" + param + "%%", parameter_value)

# write_Excel method is used to write testcase status to actual excel
def write_Excel(excelPath, SheetName, Scenario, status):
    workBook = openpyxl.load_workbook(excelPath)
    Sheet=workBook.get_sheet_by_name(name=SheetName)
    Cell_Obj = Sheet.cell(row=1, column=1)

    max_row = Sheet.max_row
    Scenario_Data = []
    if Cell_Obj.value == "Scenario":
        i = 2
        columni = 1
        while i <= max_row:
            if Scenario == Sheet.cell(row=i, column=columni).value:
                max_column = Sheet.max_column
                c1 = Sheet.cell(i, max_column)
                c1.value = status
            i += 1
    workBook.save(excelPath)

def test_portaluseralldevicesbydomain():
    url = baseUrlHeader.baseURL+"/portal/user/alldevicesbydomain"
    excelPath = "C:\\Users\\TA0134\\PycharmProjects\\API_Testing\\TestData\\UserLogin.xlsx"
    testcases = ["All valid parameter", "All valid parameter L1", "All valid parameter L2",
                 "All valid parameter L3",
                 "Blank UserId", "Blank Company Domain", "Blank Offering", "All Blank",
                 "Invalid Company Domain", "Invalid UserId", "Invalid Offering", "All Invalid"]
    testcases1 = ["All valid parameter"]
    for test in testcases:
        print("*************************" + test + "******************************")
        testData = Read_Excel(excelPath, "portalUseralldevicebydomain", test)
        i = 0
        while i < len(testData):
            if testData[i] == "Blank":
                testData[i] = testData[i].replace("Blank", "")
            i += 1

        print("Final TestData" + str(testData))
        file = open('C:\\Users\\TA0134\\PycharmProjects\\API_Testing\\Json_files\\portalUserAlldevicesbydomain.json', 'r')  # open the file in read only mode
        json_input = file.read()  # It is in string format so need to convert in json
        # Update the data in the json file
        json_input = update_content(json_input, "UserId", testData[0])
        json_input = update_content(json_input, "Domain", testData[1])
        json_input = update_content(json_input, "OfferingCategory", testData[2])
        request_json = json.loads(json_input)  # json.loads we use to convert in json format

        # Make the post request with the json input
        response = requests.post(url, request_json, headers=baseUrlHeader.headers)
        print(response.status_code)
        assert response.status_code == int(testData[3])
        if response.status_code == int(testData[3]):
            write_Excel(excelPath, "portalUseralldevicebydomain", test, "Pass:- "+response.text)
        else:
            write_Excel(excelPath, "portalUseralldevicebydomain", test, "Fail-"+response.text)

        # Get the values of the json file provided and validate the fields

        # assert response.status_code == 200
        print(response.text)  # Content on the status conde
        # print(response.headers.get("X-Powered-By")) # print a specific key value from the headers.
        print("*************************Test End******************************")
